# -*- coding: utf8 -*-
from tkinter import ttk, messagebox
from tkinter import *
import mysql.connector
import re
import matplotlib
import csv
from openpyxl import Workbook
import pathlib
import os
import pathmagic
import datetime
from matplotlib.backends.backend_tkagg import NavigationToolbar2Tk, FigureCanvasTkAgg
from matplotlib.figure import Figure
import Backend

matplotlib.use("TkAgg")


##########
def front(log1, log2):
    """Функция главной программмы"""

    def dates_listbox(listbox1, ):
        """ Формирование списка дат"""
        query = "select DATE from all_dt"
        cur.execute(query)
        datafetch = cur.fetchall()
        for i in datafetch:
            listbox1.insert(END, i)
        return listbox1

    def dates_get(nb):
        """ Извлечение дат, вызывает функцию screening dates, которая обрабатывает полученные даты"""
        dates_mass = []
        clicked_items = listbox1.curselection()
        if not clicked_items:
            messagebox.showinfo("Ошибка выбора", "Выберите дату!")
        elif len(clicked_items) > 10:
            messagebox.showinfo("Ошибка выбора", "Вы выбрали слишком много дат!")
        else:
            for item in clicked_items:
                dr1 = str(listbox1.get(item))
                datstr2 = dr1[2] + dr1[3] + dr1[4] + dr1[5] + dr1[7] + dr1[8]
                dates_mass.append(datstr2)
            screening(dates_mass, nb)
        if invalid_data:
            messagebox.showinfo("Даты которых нет", invalid_data)
        valid_data.clear()
        invalid_data.clear()

    def many_dates_before(yy, nb, h, head):
        """ Заполняет таблицу старой отчетности для выбранного банка и выбранной даты"""
        if yy[7] == '1':
            itog = '.SIM_ITOGO'
        else:
            itog = '.itog'
        query = "select sprav17.code," + yy + itog + \
                " from " + yy + " join sprav17 on" \
                                " " + yy + ".code = sprav17.code " \
                                           "where " + yy + ".REGN = " + nb
        cur.execute(query)
        text52 = cur.fetchall()
        head_text = head[4:] + '/' + head[:4]
        tree.heading("#%s" % h, text="%s" % head_text, anchor=W)
        for i in range(len(text52)):
            for j in range(1, len(folderx1)):
                if text52[i][0] == tree.item(folderx1[j])['values'][0]:
                    tree.set("Glava%s" % j, "#%s" % h, "%s" % (text52[i][1]))
        for i in range(len(text52)):
            for j in range(1, len(folderx2)):
                if text52[i][0] == tree.item(folderx2[j])['values'][0]:
                    tree.set("ABC%s" % j, "#%s" % h, "%s" % (text52[i][1]))
        for i in range(len(text52)):
            for j in range(1, len(folderx3)):
                if (text52[i][0] == tree.item(folderx3[j])['values'][0]):
                    tree.set("Razd%s" % j, "#%s" % h, "%s" % (text52[i][1]))
        for i in range(len(text52)):
            for j in range(1, len(folderx4)):
                if text52[i][0] == tree.item(folderx4[j])['values'][0]:
                    tree.set("Nums%s" % j, "#%s" % h, "%s" % (text52[i][1]))
        for i in range(len(text52)):
            for j in range(1, len(folderx5)):
                if text52[i][0] == tree.item(folderx5[j])['values'][0]:
                    tree.set("Info%s" % j, "#%s" % h, "%s" % (text52[i][1]))
        d1 = tree.item('Info261')['values'][h - 1]
        tree.set("Razd16", "#%s" % h, "%s" % d1)
        tree.set("Glava7", "#%s" % h, "%s" % d1)
        d2 = tree.item('Info263')['values'][h - 1]
        tree.set("Razd17", "#%s" % h, "%s" % d2)
        d3 = tree.item('Info265')['values'][h - 1]
        tree.set("Razd18", "#%s" % h, "%s" % d3)
        tree.set("Glava5", "#%s" % h, "%s" % 0)
        tree.set("Glava6", "#%s" % h, "%s" % 0)

    def many_dates_after(yy, nb, h, head):
        """ Заполняет таблицу новой отчетности для выбранного банка и выбранной даты"""
        query = "select sprav2.code," + yy + ".SIM_ITOGO " \
                                             "from " + yy + " join sprav2 on" \
                                                            " " + yy + ".code = sprav2.code " \
                                                                       "where " + yy + ".REGN = " + nb
        cur.execute(query)
        text52 = cur.fetchall()
        head_text = head[4:] + '/' + head[:4]
        tree.heading("#%s" % h, text="%s" % head_text, anchor=W)
        p1 = s1 = i1 = n1 = 0
        for u in range(len(text52)):
            # части 1-2, 3-4
            if int(text52[u][0]) % 10000 == 9999:
                print()
            # если с 0 в начале
            if int(text52[u][0]) == 1000:
                tree.set("Info1773", "#%s" % h, "%s" % (text52[u][1]))
            elif int(text52[u][0]) == 2000:
                tree.set("Info1774", "#%s" % h, "%s" % (text52[u][1]))
            elif int(text52[u][0]) == 3000:
                tree.set("Info1778", "#%s" % h, "%s" % (text52[u][1]))
            elif int(text52[u][0]) == 4000:
                tree.set("Info1779", "#%s" % h, "%s" % (text52[u][1]))
            elif int(text52[u][0]) == 40000:
                tree.set("Section35", "#%s" % h, "%s" % (text52[u][1]))
            elif int(text52[u][0]) == 50000:
                tree.set("Section36", "#%s" % h, "%s" % (text52[u][1]))
            # часть
            elif 0 < int(text52[u][0]) - 10000 < 5:
                p1 += 1
                tree.set("Part%s" % p1, "#%s" % h, "%s" % (text52[u][1]))
            # часть 6 и раздел
            elif int(text52[u][0]) == 61101:
                tree.set("Part6", "#%s" % h, "%s" % (text52[u][1]))
                tree.set("Section34", "#%s" % h, "%s" % (text52[u][1]))
                tree.set("Num215", "#%s" % h, "%s" % (text52[u][1]))
                tree.set("Info1780", "#%s" % h, "%s" % (text52[u][1]))
            # часть 8 и раздел
            elif int(text52[u][0]) == 81201:
                tree.set("Part8", "#%s" % h, "%s" % (text52[u][1]))
                tree.set("Section37", "#%s" % h, "%s" % (text52[u][1]))
                tree.set("Num235", "#%s" % h, "%s" % (text52[u][1]))
                tree.set("Info1876", "#%s" % h, "%s" % (text52[u][1]))
            # раздел
            elif int(text52[u][0]) % 1000 == 0:
                if s1 < 34:
                    s1 += 1
                    tree.set("Section%s" % s1, "#%s" % h, "%s" % (text52[u][1]))
            # по цифрам
            elif int(text52[u][0]) % 100 == 0 and int(text52[u][0]) % 1000 != 0:
                # if n1 < 234:
                n1 += 1
                if int(text52[u][0]) == 27100:
                    n1 += 5
                if int(text52[u][0]) == 47100:
                    n1 += 5
                if int(text52[u][0]) == 71100:
                    n1 += 3
                if int(text52[u][0]) == 71300:
                    n1 += 1
                if int(text52[u][0]) == 72300:
                    n1 += 1
                tree.set("Num%s" % n1, "#%s" % h, "%s" % (text52[u][1]))

            elif 0 < int(text52[u][0]) % 100 < 100 and \
                    int(text52[u][0]) % 10000 != 9999 and int(text52[u][0]) != 1000 and \
                    int(text52[u][0]) != 2000 and int(text52[u][0]) != 3000 and \
                    int(text52[u][0]) != 4000:

                i1 += 1
                if i1 == 1773 or i1 == 1778:
                    i1 += 2
                if i1 == 1780:
                    i1 += 1
                if i1 == 1876:
                    i1 += 1
                tree.set("Info%s" % i1, "#%s" % h, "%s" % (text52[u][1]))

        tree.set("Part5", "#%s" % h, "%s" % 0)
        tree.set("Section33", "#%s" % h, "%s" % 0)
        tree.set("Num213", "#%s" % h, "%s" % 0)
        tree.set("Num214", "#%s" % h, "%s" % 0)
        d1 = tree.item("Section35")['values'][h - 1]
        d2 = tree.item("Section36")['values'][h - 1]
        d3 = int(d1) - int(d2)
        tree.set("Part7", "#%s" % h, "%s" % d3)

    def screening(dm, nb):
        """ Обработка полученных дат. Строчка if> 201707 означает, что выбираем из новой отчетности
         Так же вызываем функцию after date или before date, которая циклом перебирает дату и отправляет её в таблицу,
         если выбранны даты старой и новой отчетности, то выдаёт ошибку"""
        md = 1
        change1 = 0
        change2 = 0
        err1 = err2 = 0
        for j in range(len(dm)):
            if int(dm[j]) > 201707:
                err1 += 1
                if err2 > 0:
                    messagebox.showinfo("Ошибка",
                                        "1 апреля 2016 г. изменилась структура формы отчета о прибылях и убытках. Для "
                                        "просмотра отчетности необходимо выбрать даты только до или после этой даты.")
                YY = after_date(dm[j], nb)
                if YY:
                    change1 += 1
                    if change1 == 1:
                        toggle1()
                    md += 1
                    many_dates_after(YY, nb, md, dm[j])

            else:
                err2 += 1
                if err1 > 0:
                    messagebox.showinfo("Ошибка",
                                        "1 апреля 2016 г. изменилась структура формы отчета о прибылях и убытках. Для "
                                        "просмотра отчетности необходимо выбрать даты только до или после этой даты.")
                YY = before_date(dm[j], nb)
                if YY:
                    change2 += 1
                    if change2 == 1:
                        toggle2()
                    md += 1
                    many_dates_before(YY, nb, md, dm[j])

    def before_date(bd, nb):
        """ см screening"""
        if bd[5] == "4":
            ds2 = '1' + bd[0] + bd[1] + bd[2] + bd[3]
        elif bd[5] == "7":
            ds2 = '2' + bd[0] + bd[1] + bd[2] + bd[3]
        elif bd[5] == "0":
            ds2 = '3' + bd[0] + bd[1] + bd[2] + bd[3]
        elif bd[5] == "1":
            di = int(bd[2] + bd[3]) - 1
            if di > 9:
                di2 = str(di)
                ds2 = '4' + bd[0] + bd[1] + di2
            else:
                di2 = str(di)
                ds2 = '4' + bd[0:2] + '0' + di2
        query1 = "select name_b from " + ds2 + 'np1' + " where regn = " + nb
        cur.execute(query1)
        b_fetch1 = cur.fetchall()
        query0 = "select name_b from " + ds2 + '_np' + " where regn = " + nb
        cur.execute(query0)
        b_fetch0 = cur.fetchall()
        if b_fetch1:
            ds3 = ds2[0:5] + "_p1"
            return ds3
        elif b_fetch0:
            ds3 = ds2[0:5] + "_p "
            return ds3
        else:
            invalid_data.append(ds2)

    def after_date(ad, nb):
        """ см screening"""
        if ad[5] == "4":
            ds2 = '1' + ad[0] + ad[1] + ad[2] + ad[3] + 'np1'
        elif ad[5] == "7":
            ds2 = '2' + ad[0] + ad[1] + ad[2] + ad[3] + 'np1'
        elif ad[5] == "0":
            ds2 = '3' + ad[0] + ad[1] + ad[2] + ad[3] + 'np1'
        elif ad[5] == "1":
            di = int(ad[2] + ad[3]) - 1
            di2 = str(di)
            ds2 = '4' + ad[0] + ad[1] + di2 + 'np1'
        query = "select name_b from " + ds2 + " where regn = " + nb
        cur.execute(query)
        after_fetch = cur.fetchall()
        if not after_fetch:
            invalid_data.append(ds2)

        else:
            ds3 = ds2[0:5] + "_p1"
            return ds3

    def banks_combobox():
        """ Формирование списка Банков"""
        query = "select regn, name_b from all_banks"
        cur.execute(query)
        bank_fetch = cur.fetchall()
        return bank_fetch

    def banks_get():
        """ Извелечение банка"""
        e = combobox1.get()
        print(e)
        #try:
        num_bank = re.findall('(\d+)', e)[0]
        dates_get(num_bank)
        #except:
        #    messagebox.showinfo("Ошибка", "Выберите банк!")

    def NEW_TREE():
        """ Функция заполняет таблицу для новой отчетности(Первый столбец)"""
        global config
        global cnx
        global cur
        global user
        global password
        user = log1[:-1]
        password = log2[:-1]

        config = {'auth_plugin': 'mysql_native_password',
                  'user': user,
                  'password': password,
                  'database': "all_in_one"}
        cnx = mysql.connector.connect(**config)
        cur = cnx.cursor(buffered=True)

        folder1 = [0] * 9
        folder2 = [0] * 38
        folder3 = [0] * 236
        folder4 = [0] * 1880
        fl1 = fl2 = fl3 = fl4 = 0
        querytree = "SELECT  name, NOM, CODE, PRSTR FROM sprav2"
        cur.execute(querytree)
        text12 = cur.fetchall()
        for i in range(len(text12)):
            if text12[i][0].find("Часть") != -1:
                fl1 += 1
                folder1[fl1] = tree.insert("", i, 'Part%s' % fl1, text="%s" % text12[i][0],
                                           values=("%s" % str(fl1) + '0000',))
                k = i + 1
                help1 = 0
                while text12[k][0].find("Часть") == -1 and k < 2414:
                    if text12[k][0].find("Раздел") != -1:
                        help1 += 1
                        fl2 += 1
                        folder2[fl2] = tree.insert(folder1[fl1], k, 'Section%s' % fl2, text="%s" % text12[k][0],
                                                   values=("%s" % str(fl1) + str(help1) + '000'))
                        l = k + 1
                        help2 = 0
                        while text12[l][0].find("Раздел") == -1 and l < 2414:
                            if text12[l][2] == '' and text12[l][0].find("Часть") == -1:
                                help2 += 1
                                fl3 += 1
                                folder3[fl3] = tree.insert(folder2[fl2], l, 'Num%s' % fl3, text="%s" % text12[l][0],
                                                           values=("%s" % str(fl1) + str(help1) + str(help2) + '00'))
                                d = l + 1
                                while text12[d][3] != 20 and d < 2414 and text12[d][3] != 120:
                                    if text12[d][3] == 10 and text12[d][2] != '' and text12[d][0].find("Часть") == -1:
                                        fl4 += 1
                                        folder4[fl4] = tree.insert(folder3[fl3], d, 'Info%s' % fl4,
                                                                   text="%s" % text12[d][0],
                                                                   values=("%s" % text12[d][2]))

                                    d += 1

                            l += 1

                    k += 1
        folder4[1877] = tree.insert(folder3[235], 2412, 'Info1877', text=" убыток после налогообложения"
                                                                         " с учетом изменений прочего совокупного "
                                                                         "дохода (символ 61102 плюс символ 81102 либо "
                                                                         "символ 61102 минус символ 81101 либо символ "
                                                                         "81102 "
                                                                         "минус символ 61101)", values="81202")

    ########
    def OLD_TREE():
        """ Функция заполняет таблицу для старой отчетности(Первый столбец)"""
        global folderx1
        global folderx2
        global folderx3
        global folderx4
        global folderx5
        folderx1 = [0] * 8
        folderx2 = [0] * 7
        folderx3 = [0] * 19
        folderx4 = [0] * 52
        folderx5 = [0] * 267
        fl1 = fl2 = fl3 = fl4 = fl5 = 0
        querytree = "SELECT  name, NOM, CODE, PRSTR FROM sprav17"
        cur.execute(querytree)
        text12 = cur.fetchall()
        for i in range(len(text12) - 11):
            if text12[i][3] == 50:
                fl1 += 1
                folderx1[fl1] = tree.insert("", i, 'Glava%s' % fl1, text="%s" % text12[i][0],
                                            values=("%s" % fl1 + '0000'),
                                            open=True)
                k = i + 1
                help1 = 0
                while text12[k][3] != 150 and k < 421:
                    if text12[k][3] == 40:
                        fl2 += 1
                        help1 += 1
                        folderx2[fl2] = tree.insert(folderx1[fl1], k, 'ABC%s' % fl2, text="%s" % text12[k][0],
                                                    values=("%s" % fl1 + '000' + str(help1)))
                        l = k + 1
                        while text12[l][3] != 140 and l < 421:
                            if text12[l][3] == 30:
                                fl3 += 1
                                folderx3[fl3] = tree.insert(folderx2[fl2], l, 'Razd%s' % fl3, text="%s" % text12[l][0],
                                                            values=("%s" % fl1 + text12[l][0][7] + '000'))
                                d = l + 1
                                help2 = 0
                                while text12[d][3] != 130 and d < 421:
                                    if text12[d][3] == 20:
                                        fl4 += 1
                                        help2 += 1
                                        folderx4[fl4] = tree.insert(folderx3[fl3], d, 'Nums%s' % fl4,
                                                                    text="%s" % text12[d][0], values=(
                                                    "%s" % fl1 + text12[l][0][7] + str(help2) + '00'))
                                        e = d + 1
                                        while text12[e][3] != 120 and e < 421:
                                            if text12[e][3] == 10:
                                                fl5 += 1
                                                folderx5[fl5] = tree.insert(folderx4[fl4], e, 'Info%s' % fl5,
                                                                            text="%s" % text12[e][0],
                                                                            values=("%s" % text12[e][2]))
                                            e += 1
                                    d += 1
                            l += 1
                    k += 1
        tree.tag_configure('RED_TAG', foreground='red', font=('Times New Roman', 9))
        folderx1[3] = tree.insert("", "end", 'Glava3', text="%s" % text12[400][0], values=1000, open=True,
                                  tag='RED_TAG')
        folderx1[4] = tree.insert("", "end", 'Glava4', text="%s" % text12[401][0], values=2000, open=True,
                                  tag='RED_TAG')
        folderx3[15] = tree.insert("ABC6", "end", 'Razd15', text="%s" % text12[402][0], values=28000)
        folderx4[49] = tree.insert("Razd15", "end", 'Nums49', text="%s" % text12[403][0], values=28001)
        folderx4[50] = tree.insert("Razd15", "end", 'Nums50', text="%s" % text12[404][0], values=28002)
        folderx4[51] = tree.insert("Razd15", "end", 'Nums51', text="%s" % text12[405][0], values=28003)
        folderx1[5] = tree.insert("", "end", 'Glava5', text="%s" % text12[406][0], values=28201, open=True,
                                  tag='RED_TAG')
        folderx1[6] = tree.insert("", "end", 'Glava6', text="%s" % text12[407][0], values=28202, open=True,
                                  tag='RED_TAG')
        folderx1[7] = tree.insert("", "end", 'Glava7', text="%s" % text12[410][0], values=30000)
        folderx3[16] = tree.insert("Glava7", "end", 'Razd16', text="%s" % text12[411][0], values=31000)
        folderx5[261] = tree.insert("Razd16", "end", 'Info261', text="%s" % text12[412][0], values=31001)
        folderx5[262] = tree.insert("Razd16", "end", 'Info262', text="%s" % text12[413][0], values=31002)
        folderx3[17] = tree.insert("Glava7", "end", 'Razd17', text="%s" % text12[414][0], values=32000)
        folderx5[263] = tree.insert("Razd17", "end", 'Info263', text="%s" % text12[415][0], values=32001)
        folderx5[264] = tree.insert("Razd17", "end", 'Info264', text="%s" % text12[416][0], values=32002)
        folderx3[18] = tree.insert("Glava7", "end", 'Razd18', text="%s" % text12[418][0], values=33000)
        folderx5[265] = tree.insert("Razd18", "end", 'Info265', text="%s" % text12[419][0], values=33001)
        folderx5[266] = tree.insert("Razd18", "end", 'Info266', text="%s" % text12[420][0], values=33002)

    #############
    def toggle1():
        """ Создаёт кнопку старой/НОВОЙ отчётности"""
        if bttnafter.config('relief')[-1] == 'sunken':
            bttnafter.config(relief="raised")
            bttnbefore.config(relief="sunken")
        else:
            bttnafter.config(relief="sunken")
            bttnbefore.config(relief="raised")
            tree.delete(*tree.get_children())
            NEW_TREE()

    def toggle2():
        """ Создаёт кнопку СТАРОЙ/новой отчётности"""
        if bttnbefore.config('relief')[-1] == 'sunken':
            bttnbefore.config(relief="raised")
            bttnafter.config(relief="sunken")
        else:
            bttnbefore.config(relief="sunken")
            bttnafter.config(relief="raised")
            tree.delete(*tree.get_children())
            OLD_TREE()

    def selectItem(a):
        """Функция передает в массивы выбранную строчку, а также заголовки этих строк,
        в которых содержится дата отчетности"""
        global cur_mas
        global cur_dat
        cur_mas = []
        cur_dat = [0] * 10
        curItem = tree.focus()
        cur_mas = tree.item(curItem)['values'][1:11]
        for cd in range(1, 11):
            cur_dat[cd - 1] = (tree.heading(cd).get('text'))

    def plot(canv, a, b):
        """Строит график """
        a = [a[i] for i in range(len(a)) if a[i] != '']
        b = [b[i] for i in range(len(b)) if b[i] != '']
        try:
            a = [datetime.datetime.strptime(d, "%m/%Y").date() for d in a]
            fig.clear()
            fig.add_subplot(111).plot(a, b)
            canv.draw()
        except:
            messagebox.showinfo("Ошибка графика", "Выберите строчку")

    #############
    def tree_to_csv():
        """Перевод значений в дереве в csv файл"""
        filename = sys.path[0] + "\\main.py"
        csv_p = filename.replace("\\", "/")
        csv3p = csv_p[:-7] + "csv3.csv"
        with open(csv3p, "w", newline='') as csv_file:
            write = csv.writer(csv_file, dialect='excel', delimiter=';')
            chld1 = tree.get_children()

            def datapar(chld):
                """Переписание строк с дерева в строкои в csv"""
                for i1 in range(len(chld)):
                    ######
                    i = 0
                    pardata_arr = []
                    #####
                    parname = tree.item(chld[i1])['text']
                    pardata = tree.item(chld[i1])['values']
                    for data in pardata:
                        pardata_arr.append(data)
                    while len(pardata) + i <= 10:
                        pardata_arr.append("")
                        i += 1
                    parametr = [parname, pardata_arr[0], pardata_arr[1], pardata_arr[2], pardata_arr[3], pardata_arr[4],
                                pardata_arr[5], pardata_arr[6], pardata_arr[7], pardata_arr[8], pardata_arr[9]]
                    write.writerow(parametr)
                    recursiv(chld[i1])

            def recursiv(element):
                """Рекурсия для получения всех строк с дерева"""
                if tree.get_children(element) == ():
                    return 1
                else:
                    chld2 = tree.get_children(element)
                    datapar(chld2)

            datapar(chld1)
            cvs2xl()

    #############
    def cvs2xl():
        """Перевод csv файла в файл excel"""
        #########
        filename = sys.path[0] + "\\main.py"
        csv_p = filename.replace("\\", "/")
        csv3p = csv_p[:-7] + "csv3.csv"
        e = combobox1.get()
        if e:
            wb = Workbook()
            ws = wb.active
            i = 1
            ########
            with open(csv3p, 'r') as f:
                for row in csv.reader(f, dialect='excel', delimiter=';'):
                    ws.append(row)
                    if 'Часть ' in row[0][0:6]:
                        ws.row_dimensions[i].hidden = True
                        ws.row_dimensions[i].outlineLevel = 1
                    elif 'Раздел' in row[0][0:6]:
                        ws.row_dimensions[i].hidden = True
                        ws.row_dimensions[i].outlineLevel = 2
                    elif '.' in row[0][1:2]:
                        ws.row_dimensions[i].hidden = True
                        ws.row_dimensions[i].outlineLevel = 3
                    else:
                        ws.row_dimensions[i].hidden = True
                        ws.row_dimensions[i].outlineLevel = 4
                    i += 1
            wb.save(csv3p[:-16] + 'Graphics/otchet.xlsx')
        else:
            """Вызываю функцию, что из дбф файлов берет значения всех банков"""
            all_banks()

    def pre():
        """Проверка отчётности данных, которая производится в Backend.py"""
        if Backend.prepare(log1, log2):
            messagebox.showinfo("Нотификация", "База данных обновлена")
        else:
            messagebox.showinfo("Нотификация", "База данных не нуждается в обновлении")

    def all_banks(dates_mass1):
        filename = sys.path[0] + "\\main.py"
        csv_p = filename.replace("\\", "/")
        for name in dates_mass1:
            # Параищоыикщпиыщшкп фигня с переводом дбф в цсв и путь найти
            dbf_p = csv_p[:-15] + "Data/"+name+".dbf"
            csv_path = Backend.dbf_to_csv(dbf_p)
            wb = Workbook()
            ws = wb.active
            prev_row = 1
            i = 1
            ws.append(["rot"])
            ########
            with open(csv_path, 'r') as f:
                for row in csv.reader(f, dialect='excel', delimiter=','):
                    if row[0] != prev_row:
                        ws = wb.create_sheet(str(row[0]))
                        i = 1
                        prev_row = row[0]
                        ws.append(row)
                    elif row[0] == prev_row:
                        ws.append(row)
                        ws.row_dimensions[i].hidden = False
                    i += 1
            wb.save(csv_p[:-15] + 'Graphics/'+ name +'.xlsx')

    def dates_allbanks():
        dates_mass1 = []  # массив с выбранными датами
        clicked_items = listbox1.curselection()
        if not clicked_items:
            messagebox.showinfo("Ошибка выбора", "Выберите дату!")
        elif len(clicked_items) > 10:
            messagebox.showinfo("Ошибка выбора", "Вы выбрали слишком много дат!")
        else:
            for item in clicked_items:
                dr1 = str(listbox1.get(item))
                bd = dr1[2] + dr1[3] + dr1[4] + dr1[5] + dr1[7] + dr1[8]
                ############
                if bd[5] == "4":
                    ds2 = '1' + bd[0] + bd[1] + bd[2] + bd[3]
                elif bd[5] == "7":
                    ds2 = '2' + bd[0] + bd[1] + bd[2] + bd[3]
                elif bd[5] == "0":
                    ds2 = '3' + bd[0] + bd[1] + bd[2] + bd[3]
                elif bd[5] == "1":
                    di = int(bd[2] + bd[3]) - 1
                    if di > 9:
                        di2 = str(di)
                        ds2 = '4' + bd[0] + bd[1] + di2
                    else:
                        di2 = str(di)
                        ds2 = '4' + bd[0:2] + '0' + di2
                ds2 = ds2 + '_p1'
                dates_mass1.append(ds2)
            print(dates_mass1)
            all_banks(dates_mass1)

    #############
    root = Tk()
    root.geometry("1323x510")
    root.resizable(False, False)
    valid_data = []
    invalid_data = []
    ########
    '''Создание пустого дерева'''
    tree = ttk.Treeview(root, selectmode='browse')
    tree["columns"] = ("two", "three", "four", "five", "six", "seven", "eight", "nine", "ten", "eleven", "twelve")
    tree.place(x=0, y=285)
    tree.column("#0", width=250, minwidth=200)
    tree.column("#1", width=70, minwidth=50, stretch=NO)
    tree.column("#2", width=100, minwidth=100)
    tree.column("#3", width=100, minwidth=100)
    tree.column("#4", width=100, minwidth=100)
    tree.column("#5", width=100, minwidth=100)
    tree.column("#6", width=100, minwidth=100)
    tree.column("#7", width=100, minwidth=100)
    tree.column("#8", width=100, minwidth=100)
    tree.column("#9", width=100, minwidth=100)
    tree.column("#10", width=100, minwidth=100)
    tree.column("#11", width=100, minwidth=100)
    tree.heading("#0", text="Наименование", anchor=W)
    tree.heading("#1", text="Символ ", anchor=CENTER)
    tree.heading("#2", text="Контрольные суммы", anchor=CENTER)
    #########
    NEW_TREE()  #
    ###########
    filename = sys.path[0] + "\\main.py"
    start = filename.replace("\\", "/")
    start = start[:-15] + "Notes/"
    ###########
    '''Menu'''
    my_menu = Menu(root)
    root.config(menu=my_menu)
    file_menu = Menu(my_menu)
    my_menu.add_cascade(label="Файл", menu=file_menu)
    null_1_menu = Menu(file_menu, tearoff=0)
    file_menu.add_command(label="О программе", command=lambda: os.system("start " + start + "rukovod.docx"))
    file_menu.add_cascade(label="Инфо", menu=null_1_menu)
    null_1_menu.add_command(label='1', command=lambda: os.system("start " + start + "1.txt"))
    null_1_menu.add_command(label='2', command=lambda: os.system("start " + start + "2.txt"))
    null_1_menu.add_command(label='3', command=lambda: os.system("start " + start + "3.txt"))
    null_1_menu.add_command(label='4', command=lambda: os.system("start " + start + "4.txt"))
    null_1_menu.add_command(label='5', command=lambda: os.system("start " + start + "5.txt"))
    file_menu.add_command(label="Проверить наличие новой отчетности", command=lambda: pre())
    file_menu.add_command(label="Выход", command=lambda: exit())
    ##########
    '''Основные детали'''
    frame = Frame(root)
    frame.place(x=300, y=0)
    dbutton = Button(frame, text="Отобразить данные банка в таблице", command=banks_get, width=34)
    dbutton.pack(side=TOP)
    plotbutton = Button(frame, text="Построить график по выбранным датам",
                        command=lambda: plot(canvas, cur_dat, cur_mas),
                        fg="blue", width=34)
    plotbutton.pack(side=TOP)
    excbutton = Button(frame, text="Выгрузить данные таблицы в excel", fg="green", width=34, command=tree_to_csv)
    excbutton.pack(side=TOP)
    vsebankibutton = Button(frame, text="Выгрузить Все банки по выбранным датам", fg="red", width=34,
                            command=dates_allbanks)
    vsebankibutton.pack(side=TOP)
    frame1 = Frame(frame)
    frame1.pack(side=TOP)
    bttnafter = Button(frame1, text="Новая отчетность", relief="sunken", width=16, command=toggle1)
    bttnafter.pack(side=LEFT)
    bttnbefore = Button(frame1, text="Старая отчетность", relief="raised", width=16, command=toggle2)
    bttnbefore.pack(side=RIGHT)
    labelwarn = Label(frame, text="1 апреля 2016 г. изменилась структура формы отчета.\n"
                                  " Для просмотра отчетности необходимо выбрать даты \nтолько до или после этой даты.")
    labelwarn.pack(side=TOP)
    value2 = StringVar()
    value2.set("Выберите Банк")
    bank_list = banks_combobox()
    combobox1 = ttk.Combobox(root, width=27, values=bank_list, height=10, font=("Calibri", 10), textvariable=value2)
    combobox1.place(x=0, y=0)
    lab = Label(root, text="Выберите даты отчётности")
    lab.place(x=0, y=100)
    listbox1 = Listbox(root, height=10, selectbackground="#708090", width=35, selectmode=MULTIPLE)
    listbox1.place(x=0, y=120)
    tree.bind('<Button-1>', selectItem)
    ch_dates = [0] * 10
    ch_num = [0] * 10
    fig = Figure(figsize=(13, 5), dpi=50)
    fig.add_subplot(111).plot(ch_dates, ch_num)
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.draw()
    canvas.get_tk_widget().place(x=700)
    dates_listbox(listbox1)
    toolbar = NavigationToolbar2Tk(canvas, root)
    toolbar.place(x=700, y=250)
    root.mainloop()
